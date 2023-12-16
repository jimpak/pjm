import openpyxl

workbook = openpyxl.load_workbook('./Excel/singer.xlsx')
wsheetList = workbook.sheetnames

for wsName in wsheetList :
    worksheet = workbook[wsName]
    print('** 워크시트의 이름 : %s' % (wsName) )
    # openpyxl은 행과 열의 시작이 1부터 진행, 마지막값을 뽑기 위해서는 1을 더해줘야함.
    for row in range(1, worksheet.max_row+1) :
        for col in range(1, worksheet.max_column+1) :
            print("%s" % (worksheet.cell(row=row, column=col).value), end='\t')
        print()
    print()