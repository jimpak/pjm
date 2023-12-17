import openpyxl
from copy import copy

workbook = openpyxl.load_workbook('./Excel/singer.xlsx')
wsheetList = workbook.sheetnames # 액셀파일의 시트이름만 리스트로 생성

outWorkbook = openpyxl.Workbook() # xlsx파일을 생성
outWorkbook.remove(outWorkbook['Sheet']) # 기본으로 생성된 시트를 일단 제거

for wsName in wsheetList :
    worksheet = workbook[wsName]
    outSheet = outWorkbook.create_sheet(wsName)
    for row in range(1, worksheet.max_row+1) :
        outSheet.row_dimensions[row].height = worksheet.row_dimensions[row].height
        for col in range(1, worksheet.max_column+1) :
            outSheet.column_dimensions[chr(ord('A')+col-1)].width \
                = worksheet.column_dimensions[chr(ord('A')+col-1)].width
            inCell = worksheet.cell(row=row, column=col)
            outCell = outSheet.cell(row=row, column=col, value= inCell.value)
            if inCell.has_style:
                outCell.font = copy(inCell.font)
                outCell.border = copy(inCell.border)
                outCell.fill = copy(inCell.fill)
                outCell.number_format = copy(inCell.number_format)
                outCell.alignment = copy(inCell.alignment)

outWorkbook.save('./Excel/singer_copy.xlsx')
print("Save. OK~")